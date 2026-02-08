#!/usr/bin/env python3
"""Generate synthetic XLSX fixture for P0.2.2 runtime truth validation.

Ground-truth counts (deterministic):
  Data sheets: 3 (Contract_A, Contract_B, Contract_C)
  Meta sheet: 1 (_change_log)
  Reference sheet: 1 (Glossary_Reference)
  Total sheets: 5

  Contract_A rows: 5 real + 1 header-echo = 6 raw (5 after sanitization)
  Contract_B rows: 4 real
  Contract_C rows: 3 real
  Orphan rows (no file_url or contract_key): 2 (in Contract_C sheet)
  _change_log rows: 3 (meta sheet - excluded from records_total)
  Glossary_Reference rows: 2 (reference sheet - excluded from records_total)

  Columns include: file_url, file_name, contract_key, record_id, status,
                    __meta_source, _glossary_ref, _system, normal_field_1, normal_field_2

  Expected after load:
    contracts_total: 3 (A, B, C - derived from file_url)
    records_total: 12 (5+4+3 data rows, meta/ref excluded)
    orphan_rows: 2 (rows without contract assignment)
    header_echo_removed: 1 (from Contract_A sheet)
    meta_sheets_excluded: 1 (_change_log)
    ref_sheets_excluded: 1 (Glossary_Reference)
    sys_fields_in_data: 3 (__meta_source, _glossary_ref, _system)
"""

from openpyxl import Workbook
import os

wb = Workbook()

HEADERS = [
    'file_url', 'file_name', 'contract_key', 'record_id',
    'status', 'document_type', '_document_role',
    '__meta_source', '_glossary_ref', '_system',
    'normal_field_1', 'normal_field_2'
]

ws1 = wb.active
ws1.title = 'Contract_A'
ws1.append(HEADERS)
for i in range(1, 6):
    ws1.append([
        'https://example.com/docs/contract-a.pdf',
        'Contract Alpha',
        'contract-a',
        f'rec_a_{i}',
        'active',
        'Root Agreement',
        'Root Agreement',
        f'source_{i}',
        f'glossary_{i}',
        f'sys_{i}',
        f'value_a_{i}',
        f'data_a_{i}'
    ])
ws1.append([
    'file_url', 'file_name', 'contract_key', 'record_id',
    'status', 'document_type', '_document_role',
    '__meta_source', '_glossary_ref', '_system',
    'normal_field_1', 'normal_field_2'
])

ws2 = wb.create_sheet('Contract_B')
ws2.append(HEADERS)
for i in range(1, 5):
    ws2.append([
        'https://example.com/docs/contract-b.pdf',
        'Contract Beta',
        'contract-b',
        f'rec_b_{i}',
        'under_review',
        'Amendment',
        'Amendment',
        f'src_b_{i}',
        '',
        '',
        f'value_b_{i}',
        f'data_b_{i}'
    ])

ws3 = wb.create_sheet('Contract_C')
ws3.append(HEADERS)
for i in range(1, 4):
    ws3.append([
        'https://example.com/docs/contract-c.pdf',
        'Contract Gamma',
        'contract-c',
        f'rec_c_{i}',
        'pending',
        'Side Letter',
        'Side Letter',
        '',
        '',
        '',
        f'value_c_{i}',
        f'data_c_{i}'
    ])
for i in range(1, 3):
    ws3.append([
        '',  # no file_url -> orphan
        '',  # no file_name
        '',  # no contract_key
        f'orphan_{i}',
        'unknown',
        '',
        '',
        '',
        '',
        '',
        f'orphan_val_{i}',
        f'orphan_data_{i}'
    ])

ws4 = wb.create_sheet('_change_log')
ws4.append(['timestamp', 'action', 'actor', 'details'])
for i in range(1, 4):
    ws4.append([f'2025-01-0{i}', 'edit', 'analyst', f'Change {i}'])

ws5 = wb.create_sheet('Glossary_Reference')
ws5.append(['term', 'definition', 'category'])
ws5.append(['Term 1', 'Definition 1', 'Category A'])
ws5.append(['Term 2', 'Definition 2', 'Category B'])

outdir = 'ui/viewer/test-data'
os.makedirs(outdir, exist_ok=True)
outpath = os.path.join(outdir, 'p022_fixture.xlsx')
wb.save(outpath)
print(f'Fixture saved to {outpath}')
print(f'Ground truth:')
print(f'  Data sheets: 3 (Contract_A=5+1echo, Contract_B=4, Contract_C=3+2orphan)')
print(f'  Meta sheets: 1 (_change_log=3 rows)')
print(f'  Ref sheets: 1 (Glossary_Reference=2 rows)')
print(f'  Expected contracts_total: 3')
print(f'  Expected records_total: 14 (5+4+3+2 data rows, header-echo kept as row but later sanitized)')
print(f'  Expected header_echo_removed: 1')
print(f'  Expected orphan_rows: 2')
print(f'  Expected sys_fields: 3 (__meta_source, _glossary_ref, _system)')
